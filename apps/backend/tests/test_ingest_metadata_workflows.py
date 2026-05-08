from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from apps.backend.tests.rag_plan_modules import *

from apps.backend.tests.rag_plan_fixtures import *


def test_ingestion_override_keeps_document_code_metadata() -> None:
    service = _build_ingestion_service_for_tests()
    detected_metadata = FileMetadata(
        document_code="RM797",
        document_code_source="filename_rule",
    )

    resolved = service._resolve_document_metadata_with_override(
        detected_metadata=detected_metadata,
        metadata_override={
            "document_code": "RM797",
            "document_code_source": "filename_rule",
        },
    )

    assert resolved.document_code == "RM797"
    assert resolved.document_code_source == "filename_rule"

def test_ingestion_without_preview_keeps_detected_document_code_only() -> None:
    service = _build_ingestion_service_for_tests()
    detected_metadata = FileMetadata(
        document_code="RM797",
        document_code_source="filename_rule",
    )

    resolved = service._resolve_document_metadata_with_override(
        detected_metadata=detected_metadata,
        metadata_override=None,
    )

    assert resolved.document_code == "RM797"
    assert resolved.document_code_source == "filename_rule"

def test_metadata_upload_requires_exact_first_file_column(tmp_path: Path) -> None:
    service = _build_metadata_upload_service(tmp_path=tmp_path)
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        "File,Id\nRM797_ID_1668,1668\n",
        encoding="utf-8",
    )

    with pytest.raises(MetadataUploadValidationError, match="first column named `file`"):
        service.parse_csv(csv_path=csv_path)

def test_metadata_upload_rejects_duplicate_files_case_insensitive(tmp_path: Path) -> None:
    service = _build_metadata_upload_service(tmp_path=tmp_path)
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        "file,Id\nRM797_ID_1668.zip,1668\nrm797_id_1668,1669\n",
        encoding="utf-8",
    )

    with pytest.raises(MetadataUploadValidationError, match="Duplicate `file` values"):
        service.upload_csv(
            user_id=7,
            csv_path=csv_path,
            source_file_name="metadata.csv",
        )

def test_metadata_upload_upload_csv_coerces_scalars_and_preserves_dynamic_fields(tmp_path: Path) -> None:
    repository = _StubMetadataRepository()
    service = _build_metadata_upload_service(
        tmp_path=tmp_path,
        repository=repository,
    )
    csv_path = tmp_path / "metadata.csv"
    csv_path.write_text(
        (
            "file,Codigo de Sitio,Id,Activo,Monto,Leading,Observacion\n"
            "RM797_ID_1668.zip,RM797,1668,si,10.5,00123,\n"
            "RM999_ID_1.pdf,RM999,1,no,0,0001,comparar\n"
        ),
        encoding="utf-8",
    )

    result = service.upload_csv(
        user_id=7,
        csv_path=csv_path,
        source_file_name="metadata.csv",
    )

    assert result.columns == [
        "file",
        "Codigo de Sitio",
        "Id",
        "Activo",
        "Monto",
        "Leading",
        "Observacion",
    ]
    assert result.total_rows == 2
    assert result.matched_files == ["RM797_ID_1668"]
    assert result.unmatched_files == ["RM999_ID_1"]
    assert len(repository.created_uploads) == 1
    assert len(repository.replaced_rows) == 1

    stored_rows = repository.replaced_rows[0]["rows"]
    assert len(stored_rows) == 2

    first_row = stored_rows[0]
    assert first_row["file_key"] == "RM797_ID_1668"
    assert first_row["search_text"] == (
        "file: RM797_ID_1668 | Codigo de Sitio: RM797 | Id: 1668 | Activo: True | "
        "Monto: 10.5 | Leading: 00123"
    )
    assert json.loads(first_row["row_json"]) == {
        "file": "RM797_ID_1668",
        "fields": {
            "Codigo de Sitio": "RM797",
            "Id": 1668,
            "Activo": True,
            "Monto": 10.5,
            "Leading": "00123",
            "Observacion": None,
        },
    }

    second_row = stored_rows[1]
    assert json.loads(second_row["row_json"]) == {
        "file": "RM999_ID_1",
        "fields": {
            "Codigo de Sitio": "RM999",
            "Id": 1,
            "Activo": False,
            "Monto": 0,
            "Leading": "0001",
            "Observacion": "comparar",
        },
    }

def test_metadata_upload_replace_csv_preserves_dataset_id_and_replaces_rows(tmp_path: Path) -> None:
    repository = _StubMetadataRepository()
    service = _build_metadata_upload_service(
        tmp_path=tmp_path,
        repository=repository,
    )
    csv_path = tmp_path / "metadata-v2.csv"
    csv_path.write_text(
        (
            "file,Id,Estado\n"
            "RM797_ID_1668,1668,Activo\n"
            "RM797_ID_5515,5515,Revision\n"
        ),
        encoding="utf-8",
    )

    result = service.replace_csv(
        metadata_upload_id="metadata-123",
        user_id=7,
        csv_path=csv_path,
        source_file_name="metadata-v2.csv",
    )

    assert result.metadata_upload_id == "metadata-123"
    assert result.columns == ["file", "Id", "Estado"]
    assert result.total_rows == 2
    assert repository.updated_uploads[0]["metadata_upload_id"] == "metadata-123"
    assert repository.replaced_rows[0]["metadata_upload_id"] == "metadata-123"
    assert [row["file_key"] for row in repository.replaced_rows[0]["rows"]] == [
        "RM797_ID_1668",
        "RM797_ID_5515",
    ]
    assert repository.refresh_calls == [{"metadata_upload_id": "metadata-123", "user_id": 7}]

def test_metadata_workbook_normalization_derives_file_column_from_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "info-docs.xlsx"
    output_path = tmp_path / "info-docs.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Prueba"
    worksheet.append(["Id", "Codigo de Sitio", "Nombre de Sitio", "Activo"])
    worksheet.append([49, "AI041", "Antena 41", True])
    worksheet.append([5515, "RM797", "Contrato Norte", False])
    workbook.save(source_path)

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.sheet_name == "Prueba"
    assert result.columns == ["file", "Id", "Codigo de Sitio", "Nombre de Sitio", "Activo"]
    assert result.total_rows == 2
    assert result.derived_file_column is True

    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Codigo de Sitio": "AI041",
            "Nombre de Sitio": "Antena 41",
            "Activo": "true",
        },
        {
            "file": "RM797_ID_5515",
            "Id": "5515",
            "Codigo de Sitio": "RM797",
            "Nombre de Sitio": "Contrato Norte",
            "Activo": "false",
        },
    ]

def test_metadata_workbook_normalization_reorders_existing_file_column(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source_path = tmp_path / "metadata.xls"
    output_path = tmp_path / "metadata.csv"

    monkeypatch.setitem(
        normalize_metadata_workbook_to_csv.__globals__,
        "_read_xls_sheet",
        lambda source_path, sheet_name=None: LoadedWorkbookSheet(
            sheet_name="Retired",
            headers=["Id", "File", "Region"],
            rows=[
                {"Id": "49", "File": "AI041_ID_49.zip", "Region": "Norte"},
            ],
        ),
    )

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.columns == ["file", "Id", "Region"]
    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Region": "Norte",
        }
    ]

def test_metadata_workbook_normalization_rejects_duplicate_derived_files(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "duplicated.xlsx"
    output_path = tmp_path / "duplicated.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Id", "Codigo de Sitio"])
    worksheet.append([49, "AI041"])
    worksheet.append([49, "AI041"])
    workbook.save(source_path)

    with pytest.raises(MetadataWorkbookNormalizationError, match="Duplicate `file` values"):
        normalize_metadata_workbook_to_csv(
            source_path=source_path,
            output_path=output_path,
        )

def test_metadata_workbook_normalization_disambiguates_duplicate_headers(tmp_path: Path) -> None:
    from openpyxl import Workbook

    source_path = tmp_path / "duplicate_headers.xlsx"
    output_path = tmp_path / "duplicate_headers.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Id", "Codigo de Sitio", "Región", "Región", "Dirección", "Dirección"])
    worksheet.append([49, "AI041", "Norte", "Norte 2", "Uno", "Dos"])
    workbook.save(source_path)

    result = normalize_metadata_workbook_to_csv(
        source_path=source_path,
        output_path=output_path,
    )

    assert result.columns == [
        "file",
        "Id",
        "Codigo de Sitio",
        "Región",
        "Región.1",
        "Dirección",
        "Dirección.1",
    ]
    with output_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows == [
        {
            "file": "AI041_ID_49",
            "Id": "49",
            "Codigo de Sitio": "AI041",
            "Región": "Norte",
            "Región.1": "Norte 2",
            "Dirección": "Uno",
            "Dirección.1": "Dos",
        }
    ]
