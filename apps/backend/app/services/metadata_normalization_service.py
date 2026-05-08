"""Normalize workbook-based metadata into canonical CSV keyed by `file`."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import unicodedata
from typing import Any

from apps.backend.app.services.metadata_keys import canonicalize_file_key, normalize_metadata_header


@dataclass(slots=True)
class LoadedWorkbookSheet:
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, str]]


@dataclass(slots=True)
class MetadataWorkbookNormalizationResult:
    source_path: Path
    output_path: Path
    sheet_name: str
    columns: list[str]
    total_rows: int
    derived_file_column: bool


class MetadataWorkbookNormalizationError(ValueError):
    """Raised when a workbook cannot be normalized into the canonical CSV contract."""


def _normalize_lookup_key(value: str | None) -> str:
    normalized = normalize_metadata_header(value)
    normalized = unicodedata.normalize("NFKD", normalized)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = " ".join(normalized.lower().split())
    return normalized


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip()


def _make_unique_headers(headers: list[str]) -> list[str]:
    used: set[str] = set()
    counters: dict[str, int] = {}
    unique_headers: list[str] = []
    for raw_header in headers:
        base_header = normalize_metadata_header(raw_header)
        if not base_header:
            unique_headers.append(base_header)
            continue
        lowered_base = base_header.lower()
        if lowered_base not in used:
            used.add(lowered_base)
            counters[lowered_base] = 1
            unique_headers.append(base_header)
            continue
        suffix = counters.get(lowered_base, 1)
        while True:
            candidate = f"{base_header}.{suffix}"
            suffix += 1
            if candidate.lower() not in used:
                used.add(candidate.lower())
                counters[lowered_base] = suffix
                unique_headers.append(candidate)
                break
    return unique_headers


def _read_xlsx_sheet(source_path: Path, *, sheet_name: str | None) -> LoadedWorkbookSheet:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MetadataWorkbookNormalizationError(
            "XLSX normalization requires `openpyxl`. Install backend requirements first."
        ) from exc

    workbook = load_workbook(filename=source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        headers: list[str] | None = None
        rows: list[dict[str, str]] = []
        for raw_row in worksheet.iter_rows(values_only=True):
            if headers is None:
                headers = _make_unique_headers(
                    [normalize_metadata_header(_cell_to_text(value)) for value in raw_row]
                )
                continue
            values = [_cell_to_text(value) for value in raw_row]
            if not any(value.strip() for value in values):
                continue
            padded = values + [""] * max(0, len(headers or []) - len(values))
            rows.append(
                {
                    header: padded[index] if index < len(padded) else ""
                    for index, header in enumerate(headers or [])
                }
            )
        if headers is None:
            raise MetadataWorkbookNormalizationError("The metadata workbook is empty.")
        return LoadedWorkbookSheet(
            sheet_name=worksheet.title,
            headers=headers,
            rows=rows,
        )
    finally:
        workbook.close()


def _read_xls_sheet(source_path: Path, *, sheet_name: str | None) -> LoadedWorkbookSheet:
    try:
        import xlrd
    except ImportError as exc:
        raise MetadataWorkbookNormalizationError(
            "XLS normalization requires `xlrd`. Install backend requirements first."
        ) from exc

    workbook = xlrd.open_workbook(source_path)
    worksheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)
    if worksheet.nrows <= 0:
        raise MetadataWorkbookNormalizationError("The metadata workbook is empty.")
    headers = _make_unique_headers(
        [normalize_metadata_header(_cell_to_text(value)) for value in worksheet.row_values(0)]
    )
    rows: list[dict[str, str]] = []
    for row_index in range(1, worksheet.nrows):
        values = [_cell_to_text(value) for value in worksheet.row_values(row_index)]
        if not any(value.strip() for value in values):
            continue
        padded = values + [""] * max(0, len(headers) - len(values))
        rows.append({header: padded[index] if index < len(padded) else "" for index, header in enumerate(headers)})
    return LoadedWorkbookSheet(
        sheet_name=worksheet.name,
        headers=headers,
        rows=rows,
    )


def _load_workbook_sheet(source_path: Path, *, sheet_name: str | None) -> LoadedWorkbookSheet:
    suffix = source_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_sheet(source_path, sheet_name=sheet_name)
    if suffix == ".xls":
        return _read_xls_sheet(source_path, sheet_name=sheet_name)
    raise MetadataWorkbookNormalizationError(
        f"Unsupported metadata workbook format `{source_path.suffix}`. Expected .xlsx or .xls."
    )


def _resolve_file_source(headers: list[str]) -> tuple[str | None, str | None, str | None]:
    header_by_key = {_normalize_lookup_key(header): header for header in headers if normalize_metadata_header(header)}
    file_header = header_by_key.get("file")
    site_header = header_by_key.get("codigo de sitio")
    id_header = header_by_key.get("id")
    return file_header, site_header, id_header


def normalize_metadata_workbook_to_csv(
    *,
    source_path: Path,
    output_path: Path,
    sheet_name: str | None = None,
) -> MetadataWorkbookNormalizationResult:
    workbook_sheet = _load_workbook_sheet(Path(source_path), sheet_name=sheet_name)
    headers = [normalize_metadata_header(header) for header in workbook_sheet.headers]
    if not headers or not any(header for header in headers):
        raise MetadataWorkbookNormalizationError("The metadata workbook must include a header row.")
    if any(not header for header in headers):
        raise MetadataWorkbookNormalizationError("The metadata workbook contains empty header names.")

    lowered_headers = [header.lower() for header in headers]
    if len(lowered_headers) != len(set(lowered_headers)):
        raise MetadataWorkbookNormalizationError("The metadata workbook contains duplicate header names.")

    file_header, site_header, id_header = _resolve_file_source(headers)
    derived_file_column = file_header is None
    if derived_file_column and not site_header:
        raise MetadataWorkbookNormalizationError(
            "The metadata workbook must include either `file` or `Codigo de Sitio` to derive it."
        )
    if derived_file_column and not id_header:
        raise MetadataWorkbookNormalizationError(
            "The metadata workbook must include either `file` or `Id` to derive it."
        )

    output_columns = ["file"] + [header for header in headers if header != file_header]
    normalized_rows: list[dict[str, str]] = []
    seen_files: set[str] = set()
    for row_number, row in enumerate(workbook_sheet.rows, start=2):
        if file_header is not None:
            raw_file = row.get(file_header, "")
        else:
            site_value = row.get(site_header or "", "")
            id_value = row.get(id_header or "", "")
            raw_file = f"{site_value}_ID_{id_value}" if site_value and id_value else ""
        file_key = canonicalize_file_key(raw_file)
        if not file_key:
            raise MetadataWorkbookNormalizationError(
                f"Row {row_number} could not produce the required `file` value."
            )
        normalized_key = file_key.lower()
        if normalized_key in seen_files:
            raise MetadataWorkbookNormalizationError(
                f"Duplicate `file` values are not allowed in normalized metadata CSV: {file_key}."
            )
        seen_files.add(normalized_key)
        normalized_row = {"file": file_key}
        for header in headers:
            if header == file_header:
                continue
            normalized_row[header] = row.get(header, "")
        normalized_rows.append(normalized_row)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=output_columns)
        writer.writeheader()
        writer.writerows(normalized_rows)

    return MetadataWorkbookNormalizationResult(
        source_path=Path(source_path),
        output_path=output_path,
        sheet_name=workbook_sheet.sheet_name,
        columns=output_columns,
        total_rows=len(normalized_rows),
        derived_file_column=derived_file_column,
    )
